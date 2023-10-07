import time
from typing import Literal
from threading import Thread

import validate_docbr as vdb
import pandas as pd
from openpyxl import Workbook, worksheet, load_workbook
from openpyxl.styles import Font, Side, Border, Alignment, NamedStyle, PatternFill

from Modulos.arquivo import AbreArquivo
from Modulos.funcoes_sanitizacao import sanitiza_cpf, sanitiza_nome
from Modulos.busca_cep import BuscaEndereco
from Modulos.imprimir import Impressao
from Modulos.models.sorteio import Sorteio, Produto
from Modulos.constants import *
from Modulos.data_hora_br import DatasBr


class Tabelas:
    def __init__(self, impressao: Impressao) -> None:
        self.__limpando = False
        self.__arquivo = AbreArquivo()
        self.__impressao = impressao
        self.__docbr = vdb.CPF()

        self.__tb_produtos: pd.DataFrame = None

        self.__tb_sorteios: pd.DataFrame = None
        self.__tb_vencedores: pd.DataFrame = None

        self.__caminho_tb_inscritos: pd.DataFrame = None
        self.__tb_inscritos: pd.DataFrame = None
        self.__tb_inscricoes_validas: pd.DataFrame = None

        self.__caminho_tb_colaboradores: pd.DataFrame = None
        self.__tb_colaboradores: pd.DataFrame = None

        self.__total_cadastros_repetidos = None
        self.__total_cpfs_invalidos = None
        self.__total_colaboradores_cadastrados = None

        self.__tb_jobs_em_andamento: pd.DataFrame = None

        self.busca = False

    def inicia_tb_jobs_em_andamento(self):
        self.__tb_jobs_em_andamento = pd.DataFrame(columns=COLUNAS_SERVICO_IMPRESSORAS)

    def update_tb_jobs_em_andamento(self, jobs_updated: list[dict]):
        del self.__tb_jobs_em_andamento
        update = pd.DataFrame(jobs_updated, columns=COLUNAS_SERVICO_IMPRESSORAS)
        self.__tb_jobs_em_andamento = update

        if not jobs_updated:
            return

        try:
            self.__tb_jobs_em_andamento.drop('Submitted', axis=1, inplace=True)
        except KeyError:
            pass

    def get_total_jobs_em_andamento(self, impressora) -> int:
        return (self.__tb_jobs_em_andamento['pPrinterName'] == impressora).sum()

    def get_log_jobs_em_andamento(self, impressora) -> str:
        df = self.__tb_jobs_em_andamento.query(f'pPrinterName=="{impressora}"')
        log = '\n'.join(df['pDocument'])
        return log

    def inicia_tb_inscritos(self, data: str, caminho=None):
        if caminho is None:
            caminho = self.__arquivo.abre_documento('Selecione o arquivo de inscritos')

        if not caminho:
            return
        df_sujo = self.__arquivo.abre_dataframe(caminho)
        df_com_data = df_sujo.assign(**{
            'Dia_do_evento': lambda df=df_sujo: df['Data'].apply(lambda x: str(x).split()[0] == data),
        })
        df_limpo = df_com_data.query('Dia_do_evento==True')

        self.__tb_inscritos = df_limpo
        self.__tb_inscricoes_validas = pd.DataFrame(columns=COLUNAS_DA_TABELA_DE_PARTICIPANTES)
        self.__caminho_tb_inscritos = caminho

    @property
    def get_total_inscritos(self):
        return self.__tb_inscritos.shape[0]

    @property
    def get_caminho_tb_inscritos(self):
        return self.__caminho_tb_inscritos

    def inicia_tb_colaboradores(self, caminho=None):
        if caminho is None:
            caminho = self.__arquivo.abre_documento('Selecione o arquivo de colaboradores')

        if not caminho:
            return

        df = self.__arquivo.abre_dataframe(caminho)
        self.__tb_colaboradores = df
        self.__caminho_tb_colaboradores = caminho

    @property
    def get_caminho_tb_colaboradores(self):
        return self.__caminho_tb_colaboradores

    def verifica_cadastros(self):
        self.__tb_inscritos[CPF] = self.__tb_inscritos[CPF].apply(sanitiza_cpf)
        self.__tb_inscritos[NOME] = self.__tb_inscritos[NOME].apply(sanitiza_nome)
        self.__tb_inscritos = self.__tb_inscritos.assign(**{
            DUPLICADA: self.__tb_inscritos.duplicated(subset=CPF, keep='last'),
            VALIDADE_CPF: lambda df: df.query(f'{DUPLICADA}==False')[CPF].apply(self.__docbr.validate),
            CPF: lambda df: df.query(f'{VALIDADE_CPF}==True')[CPF].apply(self.__docbr.mask),
            COLABORADOR: lambda df: df[CPF].isin(self.__tb_colaboradores[CPF]),
        })

        self.__total_cadastros_repetidos = (self.__tb_inscritos[DUPLICADA] == True).sum()
        self.__total_cpfs_invalidos = (self.__tb_inscritos[VALIDADE_CPF] == False).sum()
        self.__total_colaboradores_cadastrados = (self.__tb_inscritos[COLABORADOR] == True).sum()
        self.__tb_inscricoes_validas = self.__tb_inscritos.query(f'{VALIDADE_CPF}==True and {COLABORADOR}==False')

        # Thread(target=self.__busca_enderecos, daemon=True).start()

    def __busca_enderecos(self):
        self.busca = True
        infos_ceps = []
        for cep in self.__tb_inscricoes_validas.CEP:
            infos = BuscaEndereco(cep).__dict__
            infos_ceps.append(infos)
            time.sleep(2)
        df_ceps = pd.DataFrame(infos_ceps, dtype='string')
        self.__tb_inscricoes_validas = self.__tb_inscricoes_validas.merge(df_ceps, right_index=True, left_index=True)

        self.busca = False

    @property
    def get_total_cadastros_repetidos(self) -> int:
        return self.__total_cadastros_repetidos

    @property
    def get_total_cpfs_invalidos(self) -> int:
        return self.__total_cpfs_invalidos

    @property
    def get_total_colaboradores_cadastrados(self) -> int:
        return self.__total_colaboradores_cadastrados

    @property
    def get_total_inscricoes_validas(self) -> int:
        return self.__tb_inscricoes_validas.shape[0]

    @property
    def get_estado_dos_inscritos(self) -> int:
        return self.__tb_inscricoes_validas.value_counts(subset='uf')

    @property
    def get_cidades_dos_inscritos(self):
        return self.__tb_inscricoes_validas.value_counts(subset='localidade')

    @property
    def get_tb_inscricoes_validas(self):
        return self.__tb_inscricoes_validas

    @property
    def get_tb_vencedores(self):
        return self.__tb_vencedores

    def vencedor_exists(self, cpf: str):
        return any(self.__tb_inscricoes_validas[CPF] == cpf)

    def busca_vencedor(self, cpf: str):
        cpf = self.__docbr.mask(sanitiza_cpf(cpf))
        dados_vencedor: pd.DataFrame = self.__tb_inscricoes_validas.query(f'{CPF}=="{cpf}"')
        return dados_vencedor.values

    def salva_tabela(self, nome_da_tabela: Literal['tb_vencedores', 'tb_inscricoes_validas'], caminho: str):
        tabela: pd.DataFrame = getattr(self, f'get_{nome_da_tabela}')
        return self.__arquivo.salva_arquivo_filtrado(tabela, caminho, nome_da_tabela)

    def add_sorteio(self, lista_de_sorteios: list[Sorteio]):
        lista_premios = [
            [sorteio.nome_do_sorteio] + premio.valores
            for sorteio in lista_de_sorteios
            for premio in sorteio.premios
        ]
        self.__tb_sorteios = pd.DataFrame(lista_premios, columns=["sorteio"] + COLUNAS_DA_TABELA_DE_PREMIOS)

    def get_tb_sorteios_valores(self) -> [dict[list[Produto]], None]:
        sorteios = dict()
        for sorteio in self.__tb_sorteios.values:
            chave, *infos = sorteio
            sorteios.setdefault(chave, []).append(Produto(*infos))
        return sorteios

    def exportar_relatorio_vencedores(self, lista_de_sorteios: list[Sorteio]):
        wb = Workbook()
        wb.remove(wb.active)

        header_style = self.create_header_style()
        name_style = self.create_name_style()
        for sorteio in lista_de_sorteios:
            ws = wb.create_sheet(sorteio.nome_do_sorteio)
            self.set_column_widths(ws)
            self.set_header_cells(ws, sorteio, header_style, name_style)
            self.set_premio_cells(ws, sorteio)

        data_crua = lista_de_sorteios[0].data_cadastro_vencedor
        data = DatasBr(data_e_hora=data_crua)

        save_path = self.__arquivo.get_save_path(nome_base_arquivo=f'Planilha de envio {data.nome_do_mes} {data.ano}')
        if save_path is None:
            return
        wb.save(save_path)

    def create_header_style(self):
        header_style = NamedStyle('Header')
        header_style.font = Font(color=BRANCO, bold=True)
        header_style.fill = PatternFill(start_color=VERDE, end_color=VERDE, fill_type=SOLID)
        header_style.border = self.create_border()
        header_style.alignment = Alignment(horizontal=CENTER, vertical=CENTER)
        return header_style

    def create_name_style(self):
        name_style = NamedStyle('NomeParticipante')
        name_style.font = Font(color=PRETO, bold=True)
        name_style.fill = PatternFill(start_color=AMARELO, end_color=AMARELO, fill_type=SOLID)
        name_style.border = self.create_border()
        return name_style

    def create_border(self):
        linha_borda = Side(style='thin', color=PRETO)
        return Border(left=linha_borda, right=linha_borda, top=linha_borda, bottom=linha_borda, )

    def set_column_widths(self, ws: worksheet):
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        larguras = [17, 50, 17, 55, 11, 17, 22, 21]

        for col, largura in zip(columns, larguras):
            ws.column_dimensions[col].width = largura

    def set_header_cells(self, ws: worksheet, sorteio: Sorteio, header_style: NamedStyle, name_style: NamedStyle):
        ws['A1'] = sorteio.nome_do_sorteio
        ws.merge_cells('A1:B1')
        ws['A1'].style = header_style
        ws['B1'].style = header_style

        ws['A2'] = sorteio.nome_vencedor
        ws.merge_cells('A2:B2')
        ws['A2'].style = name_style
        ws['B2'].style = name_style

        header_labels = ['CPF', 'RG', 'Telefone', 'Email', 'Endereço de envio', 'CEP']
        values = [sorteio.cpf_vencedor, sorteio.rg_vencedor, sorteio.telefone_vencedor,
                  sorteio.email_vencedor, sorteio.endereco_vencedor, sorteio.cep_vencedor]
        borda = self.create_border()
        for linha, rotulo in enumerate(header_labels, start=3):
            ws.cell(row=linha, column=1, value=rotulo).border = borda
            ws.cell(row=linha, column=2, value=values[linha - 3]).border = borda

        header_row = ws['A1:H1']
        for cell in header_row[0]:
            cell.style = header_style

    def set_premio_cells(self, ws: worksheet, sorteio: Sorteio):
        ws['C1'] = 'Código do produto'
        ws['D1'] = 'Descrição'
        ws['E1'] = 'Quantidade'
        ws['F1'] = 'CC'
        ws['G1'] = 'Data fechamento do cc'
        ws['H1'] = 'Responsável pelo cc'

        premios = sorteio.__dict__['premios']

        for linha, premio in enumerate(premios, start=3):
            for coluna, valor in enumerate(premio, start=3):
                ws.cell(row=linha, column=coluna, value=valor).border = self.create_border()

    # noinspection PyTypeChecker
    def busca_produto(self, codigo: str) -> [pd.Series, None]:
        if self.__tb_produtos is None:
            self.__tb_produtos = self.__arquivo.abre_documento('Abrirtabela de produtos')
            try:
                self.__tb_produtos = pd.read_excel(
                    self.__tb_produtos, sheet_name='Consolidada', header=5, dtype='string',
                    usecols=['Código Produto', 'Descrição do Produto', 'Estado/UF/Região']
                )
            except:
                self.__tb_produtos = None
                return None

            filtro = self.__tb_produtos[self.__tb_produtos.columns[2]] == "SC"
            self.__tb_produtos = self.__tb_produtos[filtro]
            self.__tb_produtos.set_index(self.__tb_produtos.columns[0], inplace=True)

        produto = 0
        try:
            produto = self.__tb_produtos.loc[str(codigo)]
        finally:
            return produto